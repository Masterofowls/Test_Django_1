import openpyxl
from django.core.management.base import BaseCommand
from bodies.models import Product


class Command(BaseCommand):
    """Команда для импорта товаров из Excel файла"""
    
    help = 'Импортирует товары из Excel-файла'

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            help='Путь к Excel-файлу с товарами'
        )

    def handle(self, *args, **options):
        file_path = options['file_path']
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Получить заголовки из первой строки
            headers = [cell.value for cell in ws[1]]
            count = 0
            
            # Пройти по каждой строке, начиная со второй
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                
                if not data.get('Артикул'):
                    continue
                
                # update_or_create - создаст если не существует, обновит если существует
                Product.objects.update_or_create(
                    sku=data['Артикул'],
                    defaults={
                        'name':        data.get('Наименование товара', ''),
                        'price':       str(data.get('Цена', 0)).replace(',', '.'),
                        'description': data.get('Описание товара', '') or '',
                    }
                )
                count += 1
            
            self.stdout.write(
                self.style.SUCCESS(f'✅ Успешно импортировано {count} товаров')
            )
            
        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(f'❌ Файл не найден: {file_path}')
            )
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'❌ Ошибка при импорте: {str(e)}')
            )
