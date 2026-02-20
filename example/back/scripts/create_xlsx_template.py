#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Скрипт для создания шаблона XLSX файла для импорта товаров
"""

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Ошибка: установите openpyxl командой: pip install openpyxl")
    exit(1)

def create_xlsx_template():
    """Создать шаблон XLSX файла"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Товары"
    
    # Заголовки
    headers = ["Название", "Артикул", "Цена", "Описание"]
    
    # Форматирование заголовков
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Примеры данных
    sample_data = [
        ["Ноутбук Dell XPS", "DELL-XPS-001", 45000, "Мощный ноутбук с процессором Intel i7"],
        ["Смартфон Samsung Galaxy", "SAM-GAL-001", 25000, "Смартфон с экраном 120Hz"],
        ["Наушники Sony WH-1000", "SONY-WH-001", 8000, "Беспроводные наушники с шумоподавлением"],
        ["Монитор LG UltraWide", "LG-ULT-001", 15000, "27 дюймов UltraWide 4K монитор"],
        ["Клавиатура Corsair", "CORS-MECH-001", 5000, "Механическая клавиатура для геймеров"],
    ]
    
    for row_num, row_data in enumerate(sample_data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Установка ширины столбцов
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 12
    worksheet.column_dimensions['D'].width = 40
    
    # Морозка первой строки
    worksheet.freeze_panes = "A2"
    
    # Сохранение
    workbook.save('products_template.xlsx')
    print("✓ Шаблон создан: products_template.xlsx")


if __name__ == '__main__':
    create_xlsx_template()
